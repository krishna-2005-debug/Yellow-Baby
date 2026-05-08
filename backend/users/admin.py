"""Upgraded Users Admin — CSV export, role badges, address count."""

import csv
import openpyxl
from openpyxl.styles import Font, PatternFill
from django.contrib.admin import ModelAdmin, action
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.http import HttpResponse
from django.utils.html import format_html
from django.utils.safestring import mark_safe
from yellow_baby_backend.admin_site import yellow_baby_admin
from .models import User, OTP, Address


class UserAdmin(BaseUserAdmin):
    list_display = [
        'mobile', 'name', 'email', 'role_badge',
        'address_count', 'order_count', 'is_active', 'created_at',
    ]
    list_filter = ['role', 'is_active']
    search_fields = ['mobile', 'name', 'email']
    ordering = ['-created_at']
    readonly_fields = ['created_at']
    list_per_page = 25
    actions = ['export_users_excel', 'export_users_csv', 'activate_users', 'deactivate_users']

    fieldsets = (
        (None, {'fields': ('mobile', 'password')}),
        ('Personal Info', {'fields': ('name', 'email')}),
        ('Permissions', {'fields': ('role', 'is_active', 'is_staff', 'is_superuser', 'groups', 'user_permissions')}),
        ('Dates', {'fields': ('created_at',)}),
    )
    add_fieldsets = (
        (None, {
            'classes': ('wide',),
            'fields': ('mobile', 'name', 'password1', 'password2', 'role'),
        }),
    )

    def role_badge(self, obj):
        if obj.role == 'admin':
            return mark_safe(
                '<span style="background:#EDE9FE;color:#5B21B6;padding:3px 10px;'
                'border-radius:12px;font-size:11px;font-weight:700;">👑 Admin</span>'
            )
        return mark_safe(
            '<span style="background:#D1FAE5;color:#065F46;padding:3px 10px;'
            'border-radius:12px;font-size:11px;font-weight:700;">👤 Customer</span>'
        )
    role_badge.short_description = 'Role'

    def address_count(self, obj):
        c = obj.addresses.count()
        return f'{c} address{"es" if c != 1 else ""}'
    address_count.short_description = 'Addresses'

    def order_count(self, obj):
        c = obj.orders.count()
        return format_html('<strong>{}</strong> order{}', c, 's' if c != 1 else '')
    order_count.short_description = 'Orders'

    @action(description='⬇️ Export users to CSV')
    def export_users_csv(self, request, queryset):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="yellow_baby_users.csv"'
        writer = csv.writer(response)
        writer.writerow(['ID', 'Mobile', 'Name', 'Email', 'Role', 'Active', 'Joined'])
        for u in queryset:
            writer.writerow([
                u.id, u.mobile, u.name, u.email,
                u.role, 'Yes' if u.is_active else 'No',
                u.created_at.strftime('%d-%m-%Y'),
            ])
        return response

    @action(description='📊 Export users to Excel (.xlsx)')
    def export_users_excel(self, request, queryset):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="yellow_baby_customers.xlsx"'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Customers"
        
        headers = ['ID', 'Mobile', 'Name', 'Email', 'Role', 'Active', 'Joined', 'Total Orders', 'Total Addresses']
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            
        # Data rows
        for u in queryset:
            ws.append([
                u.id, 
                str(u.mobile), 
                str(u.name or ''), 
                str(u.email or ''),
                str(u.role).title(), 
                'Yes' if u.is_active else 'No',
                u.created_at.strftime('%Y-%m-%d %H:%M'), 
                u.orders.count(),
                u.addresses.count()
            ])
            
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    val_str = str(cell.value)
                    if len(val_str) > max_length:
                        max_length = len(val_str)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = min(adjusted_width, 50)
            
        wb.save(response)
        return response

    @action(description='✅ Activate selected users')
    def activate_users(self, request, queryset):
        updated = queryset.update(is_active=True)
        self.message_user(request, f'{updated} user(s) activated.')

    @action(description='🚫 Deactivate selected users')
    def deactivate_users(self, request, queryset):
        updated = queryset.update(is_active=False)
        self.message_user(request, f'{updated} user(s) deactivated.')


class OTPAdmin(ModelAdmin):
    list_display = ['mobile', 'code', 'is_used', 'is_valid_display', 'created_at']
    list_filter = ['is_used']
    search_fields = ['mobile']
    readonly_fields = ['created_at']

    def is_valid_display(self, obj):
        if obj.is_valid():
            return mark_safe('<span style="color:green;font-weight:bold;">&#10003; Valid</span>')
        return mark_safe('<span style="color:red;">&#10007; Expired</span>')
    is_valid_display.short_description = 'Status'


class AddressAdmin(ModelAdmin):
    list_display = ['name', 'user_mobile', 'city', 'state', 'pincode', 'is_default']
    list_filter = ['state', 'is_default']
    search_fields = ['name', 'user__mobile', 'city', 'pincode']

    def user_mobile(self, obj):
        return obj.user.mobile
    user_mobile.short_description = 'Customer Mobile'


yellow_baby_admin.register(User, UserAdmin)
yellow_baby_admin.register(OTP, OTPAdmin)
yellow_baby_admin.register(Address, AddressAdmin)
